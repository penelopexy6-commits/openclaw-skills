#!/usr/bin/env python3
"""
Step 1.2 — 商品详情API → 逐条写入Sheet 3
用法: python3 step12_product_detail.py
"""
import requests, openpyxl, time, sys, os
from openpyxl.styles import Font
from datetime import datetime, timezone, timedelta

sys.stdout.reconfigure(line_buffering=True)

# ===== 配置 =====
API_BASE = "http://api.seerfar.cn"
API_KEY = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJzaWduIjoiOGMwMWVmMGI0YjI4NDU5MzgwZTg5Mjc3OTkxYjAwMjgiLCJwYXJ0bmVyVHlwZSI6MCwicGFydG5lcklkIjo5NTEwOX0.pOup3WgVs_fFZweLExP5ck1M3EJkf2uMa7jgYEhUcVk"
import sys
EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else r"{WORKSPACE}\Step1_skus_filtered.xlsx"
LOG_PATH = r"{WORKSPACE}\step12_progress.log"
HEADERS = {
    "Authorization": f"Bearer {API_KEY}",
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Content-Type": "application/json",
    "User-Agent": "PostmanRuntime-ApipostRuntime/1.1.0"
}
MOSCOW_TZ = timezone(timedelta(hours=3))
RATE_LIMIT = 5  # seconds

# ===== Sheet 3 列映射 =====
COL_SKU = 1
COL_PLATFORM = 2
COL_URL = 3
COL_TITLE = 4
COL_CATEGORY = 5
COL_URL2 = 6
COL_IMG = 7
COL_BRAND = 8
COL_CAT2 = 9
COL_PRICE = 10
COL_DISCOUNT = 11
COL_PROMO = 12
COL_MARGIN = 13
COL_SALES30 = 14
COL_DAILY_REV = 15
COL_DAILY_SALES = 16
COL_DRR = 17
COL_SESSION = 18
COL_SEARCH_SESSION = 19
COL_CONV_CART_SEARCH = 20
COL_CONV_CART_PDP = 21
COL_CONV_VIEW_ORDER = 22
COL_CONV_ORDER = 23
COL_RETURN_RATE = 24
COL_AD = 25
COL_UPTIME = 26
COL_IMG2 = 27
COL_SELLER_TYPE = 28
COL_FULFILLMENT = 29
COL_WEIGHT = 30
COL_VOLUME = 31
COL_DIMENSION = 32
COL_CROSSBORDER = 33

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")

def read_skus(excel_path):
    """从 Sheet 2 读取所有SKU"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['SKU']
    skus = []
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[1] and isinstance(row[1], (int, float)):
            skus.append(int(row[1]))
    wb.close()
    return skus

def get_existing_skus(excel_path):
    """获取 Sheet 3 已有的SKU（避免重复写）"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['shuju']
    existing = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            existing.add(int(row[0]))
    wb.close()
    return existing

def api_product_detail(sku):
    """调SeerFar商品详情API，返回 data 或 None"""
    for attempt in range(3):
        try:
            r = requests.post(
                f"{API_BASE}/open-api/product/detail/search/ozon",
                json={"sku": str(sku), "dateRange": "past_30_days"},
                headers=HEADERS, timeout=30
            )
            if r.status_code == 429:
                wait = 15 * (attempt + 1)
                log(f"  429, wait {wait}s (attempt {attempt+1})...")
                time.sleep(wait)
                continue
            r.raise_for_status()
            data = r.json()
            if data.get("code") == 200:
                return data.get("data")
            else:
                log(f"  API error: {data.get('msg', data)}")
                return None
        except Exception as e:
            log(f"  Exception: {e}")
            if attempt < 2:
                time.sleep(10)
    return None

def write_product_row(wb, sku, data):
    """写入一条商品数据到 Sheet 3"""
    ws = wb['shuju']
    # 找下一空行
    next_row = ws.max_row + 1
    if next_row == 1:
        # 没有表头才创建
        pass

    p = data.get("product", {})
    
    # 安全取值
    def sd(d, *keys, default=""):
        v = d
        for k in keys:
            if isinstance(v, dict):
                v = v.get(k)
            else:
                return default
        return v if v is not None else default

    # 图片
    img_list = sd(p, "imageUrls")
    if isinstance(img_list, list):
        first_img = img_list[0] if img_list else ""
    else:
        first_img = ""

    # 上架时间
    up_time = sd(p, "upTime")
    if up_time and isinstance(up_time, (int, float)):
        try:
            up_str = datetime.fromtimestamp(up_time / 1000, tz=MOSCOW_TZ).strftime("%Y-%m-%d")
        except:
            up_str = str(up_time)
    else:
        up_str = ""

    # 日均销售额
    total_rev = sd(data, "totalRevenue", default=0)
    if total_rev:
        daily_rev = round(total_rev / 30, 2)
    else:
        daily_rev = 0

    # 仓储模式
    ful_list = sd(p, "fulfillment")
    if isinstance(ful_list, list):
        ful_str = ful_list[0] if ful_list else ""
    else:
        ful_str = ""

    vals = [
        sku,                                                                # 1 SKU
        "OZON",                                                             # 2 平台
        sd(p, "productUrl"),                                                # 3 链接
        sd(p, "title"),                                                     # 4 标题
        sd(p, "categoryInfo", "cnTitlePath"),                               # 5 类目
        sd(p, "productUrl"),                                                # 6 链接(重复)
        first_img,                                                          # 7 图片URL
        sd(p, "brandName"),                                                 # 8 品牌名称
        sd(p, "categoryInfo", "cnTitlePath"),                               # 9 类目(重复)
        sd(p, "price"),                                                     # 10 售价
        sd(p, "discount"),                                                  # 11 折扣率
        sd(p, "promoRevenueShare"),                                         # 12 促销收入占比
        sd(p, "grossMargin"),                                               # 13 毛利率
        sd(data, "totalSales"),                                             # 14 近30天总销量
        daily_rev,                                                          # 15 日均销售额
        sd(data, "dailySales"),                                             # 16 日均销量
        sd(p, "drr"),                                                       # 17 广告费用份额
        sd(p, "sessionCount"),                                              # 18 总访问量
        sd(p, "sessionCountSearch"),                                        # 19 搜索流量
        sd(p, "convToCartSearch"),                                          # 20 搜索→加购转化率
        sd(p, "convToCartPdp"),                                             # 21 详情→加购转化率
        sd(p, "convViewToOrder"),                                           # 22 浏览→下单转化率
        sd(p, "orderConversionRate"),                                       # 23 下单转化率
        sd(p, "returnCancellationRate"),                                    # 24 退货取消率
        sd(p, "drr"),                                                       # 25 广告
        up_str,                                                             # 26 上架时间
        first_img,                                                          # 27 商品图片链接
        sd(p, "sellerType"),                                                # 28 卖家类型
        ful_str,                                                            # 29 仓储模式
        sd(p, "weight"),                                                    # 30 重量(g)
        sd(p, "volume"),                                                    # 31 体积(L)
        sd(p, "dimension"),                                                 # 32 包装尺寸(mm)
        sd(p, "categoryInfo", "crossBorderSellable"),                       # 33 是否允许跨境
    ]

    for col, val in enumerate(vals, 1):
        ws.cell(row=next_row, column=col, value=val)

def ensure_header(excel_path):
    """确保 Sheet 3 有表头"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['shuju']
    if ws.max_row == 0 or ws.cell(1, 1).value is None:
        headers = [
            "SKU","平台","链接","标题","类目","链接(重复)","图片URL","品牌名称",
            "类目(重复)","售价(руб)","折扣率","促销收入占比","毛利率",
            "近30天总销量","日均销售额(руб)","日均销量","广告费用份额(%)",
            "近30天总访问量","近30天搜索流量","搜索加购转化率(%)",
            "详情加购转化率(%)","浏览下单转化率(%)","下单转化率(%)",
            "退货取消率(%)","广告","上架时间","商品图片链接",
            "卖家类型(0本土/1跨境)","仓储模式","重量(g)","体积(L)","包装尺寸(mm)","允许跨境"
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h).font = Font(bold=True)
    wb.save(excel_path)
    wb.close()

# ===== Main =====
def main():
    log("=" * 50)
    log("Step 1.2 — 商品详情API → 逐条写入Sheet 3")
    log("=" * 50)

    # 确保表头
    ensure_header(EXCEL_PATH)

    # 读取SKU
    all_skus = read_skus(EXCEL_PATH)
    log(f"Sheet 2 共读取 {len(all_skus)} 个SKU")

    # 获取已存在的
    existing = get_existing_skus(EXCEL_PATH)
    log(f"Sheet 3 已有 {len(existing)} 个SKU数据")

    # 过滤掉已存在的
    todo = [s for s in all_skus if s not in existing]
    log(f"待处理: {len(todo)} 个SKU")

    if not todo:
        log("全部已完成，无需处理")
        return

    success = 0
    failed = 0
    total = len(todo)

    for i, sku in enumerate(todo, 1):
        log(f"[{i}/{total}] SKU {sku}...")
        
        data = api_product_detail(sku)
        
        if data:
            try:
                wb = openpyxl.load_workbook(EXCEL_PATH)
                write_product_row(wb, sku, data)
                wb.save(EXCEL_PATH)
                wb.close()
                p = data.get("product", {})
                price = p.get("price", "?")
                title = (p.get("title", "") or "")[:40]
                log(f"  OK — price={price}, {title}")
                success += 1
            except Exception as e:
                log(f"  Write error: {e}")
                failed += 1
        else:
            log(f"  FAILED (API returned no data)")
            failed += 1

        # 限频等待（最后一次不用等）
        if i < total:
            time.sleep(RATE_LIMIT)

    # 最终汇总
    remaining = total - success - failed
    log("-" * 40)
    log(f"完成! 成功: {success}, 失败: {failed}, 跳过: {remaining}")
    log(f"文件: {EXCEL_PATH}")

    # 查积分
    try:
        r = requests.get(f"{API_BASE}/open-api/quota", headers=HEADERS, timeout=10)
        if r.status_code == 200:
            d = r.json().get("data", {})
            used = d.get("creditUsed", 0)
            limit = d.get("creditLimit", 0)
            log(f"积分: {limit - used}/{limit} (已用: {used})")
    except:
        pass

if __name__ == "__main__":
    main()
