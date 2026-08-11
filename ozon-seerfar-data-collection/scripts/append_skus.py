#!/usr/bin/env python3
"""将搜到的SKU写入 Excel SKU Sheet"""
import openpyxl, sys, json
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"{WORKSPACE}\20260722.xlsx"

if len(sys.argv) < 3:
    print("用法: python3 append_skus.py <品类名> <sku1,sku2,...>")
    sys.exit(1)

category = sys.argv[1]
skus = sys.argv[2].split(',')

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['SKU']

# 找到末尾
next_row = ws.max_row + 1

# 写品类头
ws.cell(next_row, 1, category)
next_row += 1

# 写SKU
for i, sku in enumerate(skus):
    sku = sku.strip()
    if sku:
        ws.cell(next_row + i, 2, int(sku))

wb.save(EXCEL_PATH)
wb.close()
print(f"OK: {category} → {len(skus)}个SKU写入")
